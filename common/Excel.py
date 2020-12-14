#!/usr/bin/python3
# _*_ coding:utf-8 _*_
# @Author:啊志
# @Time:2019/9/23 9:20
# @Email:1427245437@qq.com
#excel类  读内容  写内容
import  openpyxl
class excel():
    def __init__(self,file,tablename):
        self.file=openpyxl.load_workbook(file)
        self.tablename=self.file[tablename]
    def __del__(self):
        self.file.close()

    def red_excel(self,hang,lei):
        #读具体单元格中的内容
        cell_value=self.tablename.cell(row=hang,column=lei).value
        return  cell_value

    def reds_excel(self):
        max_rows=self.tablename.max_row
        max_columns=self.tablename.max_column

        all_data=[]
        for i in range(1,max_rows+1):
            rows = []
            for j in range(1,max_columns+1):
                rows.append(self.tablename.cell(row=i, column=j).value)

            all_data.append(tuple(rows))

        return all_data

    #读取任意行
    def red_any_row(self,number):
        pass

    #读取任意列
    def red_any_column(self,number):
        pass