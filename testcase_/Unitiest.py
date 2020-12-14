#!/usr/bin/python3
# _*_ coding:utf-8 _*_
# @Author:啊志
# @Time:2019/9/23 9:20
# @Email:1427245437@qq.com
import openpyxl
import unittest
from day7 import Login
class Logintest(unittest.TestCase):
    def __init__(self,hang,name,passwd,expect):
        super().__init__("test_login")
        self.name=name
        self.passwd=passwd
        self.expect=eval(expect)
        self.hang=int(hang)+1

    def setUp(self):
        self.wob=openpyxl.load_workbook("2.xlsx")
        self.bd=self.wob['Sheet1']
        print("执行用例前都会执行这个方法，类似前置条件")

    def tearDown(self):
        self.wob.save("2.xlsx")
        self.wob.close()
        print("执行每条用例后，再来执行我")

    def test_login(self):

        rel=Login.login_check(self.name,self.passwd)
        print('预期----',self.expect)
        print('实际----',rel)
        max_column=self.bd.max_column
        try:
            self.assertEqual(rel,self.expect)
            # result="成功"
            self.bd.cell(row=self.hang,column=max_column,value="成功")
        except  AssertionError:
            self.bd.cell(row=self.hang,column=max_column,value="失败")
            print("测试不通过，预期结果与实际结果不同")
if __name__=="__main__":
    unittest.main()
