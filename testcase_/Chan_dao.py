#!/usr/bin/python3
# _*_ coding:utf-8 _*_
from day8 import Excel
from ddt import ddt,data,unpack
import openpyxl
from test_unittest_demo.common import logging_template
from zi_dong_hua import Deng_lu_1
import unittest
log=logging_template.MyLogging()
a=Excel.excel(r"D:\python\python代码\test_unittest_demo\data\chandao.xlsx","Sheet")
shu_ju=a.reds_excel()
shu_ju.pop(0)
# print(shu_ju)
new_shu_ju=[]
for i in shu_ju:
    l=[]
    l.append(int(i[1]))
    l.append(i[2])
    l.append(str(i[3]))
    l.append(eval(i[4]))
    new_shu_ju.append(tuple(l))
print(new_shu_ju)
# for i in new_shu_ju:
#     uname=i[0]
#     passwd=i[1]
#     Deng_lu_1.deng_lu(uname,passwd)
@ddt
class chan_dao_test(unittest.TestCase):
    def setUp(self) :
        self.wod = openpyxl.load_workbook(r"D:\python\python代码\test_unittest_demo\data\chandao.xlsx")
        self.bd = self.wod["Sheet"]

    def tearDown(self):
        self.wod.save(r"D:\python\python代码\test_unittest_demo\data\chandao.xlsx")
        self.wod.close()

    @data(*new_shu_ju)
    @unpack
    def test_chan_dao(self,hang,uname,passwd,expect):
        rel=Deng_lu_1.deng_lu(uname,passwd)
        print('实际----', rel)
        print('预期----', expect)

        # max_column = self.bd.max_column
        # try:
        #     self.assertEqual(rel, expect)
        #     jie_guo = "成功"
        #     self.bd.cell(row=hang+1, column=max_column, value=jie_guo)
        #     log.info('通过')
        # except AssertionError:
        #     jie_guo = "失败"
        #     self.bd.cell(row=hang+1, column=max_column, value=jie_guo)
        #     log.error("失败")
        #     print('预期与实际不同，测试失败')
if __name__ == "__main__":
    unittest.main()