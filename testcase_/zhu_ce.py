#!/usr/bin/python3
# _*_ coding:utf-8 _*_
# @Author:啊志
# @Time:2019/9/23 9:20
# @Email:1427245437@qq.com
import openpyxl
import unittest
from day7 import Login
from test_unittest_demo.common import logging_template
from day8 import Excel   #数据读取
from test_unittest_demo.common import excel_reading #读取指定位置数据
from ddt import ddt,data,unpack
log=logging_template.MyLogging()
# a=Excel.excel(r"D:\python\python代码\test_unittest_demo\data\zhu_ce.xlsx","Sheet1")
# shu_ju=a.reds_excel()
# log.info("数据读取")
# # print(shu_ju)
# shu_ju.pop(0)
# new_shu_ju=[]
# for i in shu_ju:
#     l=[]
#     l.append(i[1])
#     l.append(i[2])
#     l.append(str(i[3]))
#     l.append(str(i[4]))
#     l.append(eval(i[5]))
#     new_shu_ju.append(tuple(l))
# print(new_shu_ju)

a=excel_reading.Excel(r"D:\python\python代码\test_unittest_demo\data\zhu_ce.xlsx","Sheet1")
shu_ju=a.row_col_appointed([2,3,4,5,6],[2,3,4,5,6])
# print(eval(shu_ju[0][3]))
print(shu_ju)
new_shu_ju=[]
for i in shu_ju:
    a=(list(i))

    a[4]=(eval(a[4]))
    new_shu_ju.append(tuple(a))
print(new_shu_ju)




@ddt
class zhu_ce_ddt_test(unittest.TestCase):
    def setUp(self):
        self.wod=openpyxl.load_workbook(r"D:\python\python代码\test_unittest_demo\data\zhu_ce.xlsx")
        self.bd=self.wod["Sheet1"]
    def tearDown(self):
        self.wod.save(r"D:\python\python代码\test_unittest_demo\data\zhu_ce.xlsx")
        self.wod.close()
    @data(*new_shu_ju)
    @unpack
    def test_zhu_ce(self,hang,uname,passwd,Confirm_password,expect):
        rel=Login.register_check(uname,passwd,Confirm_password)
        print('实际----', rel)
        print('预期----', expect)

        max_column = self.bd.max_column
        try:
            self.assertEqual(rel,expect)
            jie_guo="成功"
            self.bd.cell(row=int(hang)+1,column=max_column,value=jie_guo)
            log.info('通过')
        except AssertionError:
            jie_guo = "失败"
            self.bd.cell(row=int(hang)+1,column=max_column, value=jie_guo)
            log.error("失败")
            print('预期与实际不同，测试失败')
if __name__ == "__main__":
    unittest.main()
