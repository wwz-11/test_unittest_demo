#!/usr/bin/python3
# _*_ coding:utf-8 _*_
# @Author:啊志
# @Time:2019/9/23 9:20
# @Email:1427245437@qq.com

from HTMLTestRunnerNew import HTMLTestRunner

import openpyxl
import unittest
from deng_lu_dan_yuan_ce_shi import Unitiest
wob=openpyxl.load_workbook("2.xlsx")
bd=wob["Sheet1"]
#获取总行数
max_row=bd.max_row
# print(max_row)
#获取总列数
max_column=bd.max_column
# print(max_column)

hang_list=[]
zong_hang_list=[]
for i in range(2,max_row+1):
    for j in range(2,max_column):
        zhi=bd.cell(row=i,column=j).value

        hang_list.append(str(zhi))
    zong_hang_list.append(tuple(hang_list))
    hang_list.clear()

# print(zong_hang_list)
suite=unittest.TestSuite()
for i in zong_hang_list:
    # print(i)
    suite.addTest(Unitiest.Logintest(*i))
with open('1.html','wb' ) as f:
    test=HTMLTestRunner(
                        stream=f,
                       verbosity=4,                #详细程度
                       title='我的第一份测试报告',
                       description="第一份测试报告",
                       tester="啊志")
    test.run(suite)

